#APPROACH 2
#user enters the ctc
#loop through columnA of salary structure  and checks whether there is such a column which is the most independent one. 
#if found, change the value of column to the entered ctc
#call the functionsvone by one which is having formulas to update the data in cells which are dependent on this ctc

#LIMITATION:this method wont work with a differnt salary structure.

import openpyxl
 
wb = openpyxl.load_workbook("test.xlsx",data_only=True)
 
ws = wb.active
a=0
basic_salaryy=0

#calculating basic salary based on ctc entered
def basic_salary():
    for x in range(9,20):
        if ws['A'+str(x)].value == 'basic_salary':
            ws['B'+str(x)].value=ws['B'+str(a)].value*0.5
            ws['C'+str(x)].value=(ws['B'+str(a)].value*0.5)/12
            wb.save("test.xlsx")
            basic_salaryy=ws['C'+str(x)].value
            print("monthly basic salary =",basic_salaryy)
            print("annual basic salary =", ws['B'+str(x)].value)

#calculating hra based on ctc entered
def hra():
    for x in range(9,20):
        if ws['A'+str(x)].value == 'HRA':
            ws['B'+str(x)].value=ws['B'+str(x-1)].value*0.5
            ws['C'+str(x)].value=(ws['B'+str(x)].value)/12
            wb.save("test.xlsx")
            monthly_hra=ws['C'+str(x)].value
            annual_hra=ws['B'+str(x)].value
            print("annual hra=",annual_hra)
            print("monthly hra=",monthly_hra)
            
#calculating esi based on ctc entered
def employer_esi():
    for x in range(9,20):
        if ws['A'+str(x)].value == 'Employer_ESI':
            if ws['B'+str(x+1)].value<252001:
                ws['B'+str(x)].value=ws['B'+str(x+1)].value*0.0325
                ws['C'+str(x)].value=(ws['B'+str(x)].value)/12
            else:
                ws['B'+str(x)].value=0
                ws['C'+str(x)].value=(ws['B'+str(x)].value)/12
            wb.save("test.xlsx")
            monthly_ESI=ws['C'+str(x)].value
            annual_ESI=ws['B'+str(x)].value
            print("annual ESI=",annual_ESI)
            print("monthly ESI=",monthly_ESI)

#calculating pf based on ctc entered
def pf():
    for x in range(9,20):
        if ws['A'+str(x)].value == 'PF':
            ws['C'+str(x)].value=1800
            ws['B'+str(x)].value=ws['C'+str(x)].value*12
            
            wb.save("test.xlsx")
            monthly_pf=ws['C'+str(x)].value
            annual_pf=ws['B'+str(x)].value
            print("annual pf=",annual_pf)
            print("monthly pf=",monthly_pf)

#calculating special allowance  based on ctc entered
def sp_allowances():
    for x in range(9,20):
        if ws['A'+str(x)].value == 'Special Allowance':
            ws['B'+str(x)].value=ws['B'+str(a)].value-ws['B'+str(x-2)].value-ws['B'+str(x-1)].value-ws['B'+str(x+1)].value-ws['B'+str(x+2)].value
            ws['C'+str(x)].value=(ws['B'+str(x)].value)/12
            
            wb.save("test.xlsx")
            monthly_allowances=ws['C'+str(x)].value
            annual_allowances=ws['B'+str(x)].value
            print("annual allowances=",annual_allowances)
            print("monthly allowances=",monthly_allowances)


base_salary=(int(input("enter ctc")))
#looping for finding column corresponds to enterd ctc
for x in range(9,20):
    if ws['A'+str(x)].value == 'base_salary':
        ws['B'+str(x)].value=base_salary
        a=x
        

wb.save("test.xlsx")

basic_salary()
hra()
employer_esi()
pf()
sp_allowances()
