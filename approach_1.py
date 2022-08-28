#APPROACH 1
#user enters the ctc
#loop through columnA and checks whether there is such a column which is the most independent one. 
#if found, change the value of column to the entered ctc
#get the expressions for each cells using .value
#update every cells whose expression contains the independent value
#update every cells which is depending the updated cells
import openpyxl
 
wb = openpyxl.load_workbook("test.xlsx",data_only=True)
 
ws = wb.active

a=0
base_salary=(int(input("enter ctc")))
for x in range(9,20):
    if ws['A'+str(x)].value == 'base_salary':
        ws['B'+str(x)].value=base_salary
        a=x
print(a,"base")
wb.save("test.xlsx")
# base_salary=ws['B14'].value

# for x in range(9,14):
#     for y in range(1,3):
#         char = chr(65+y)
#         temp=char+str(x)
#         if 'B14' in ws[temp].value:
#             print('yes')



# for x in range(9,14):
#     for y in range(1,3):
#         char = chr(65+y)
#         temp=char+str(x)
#         alg= ws[temp].value
#         print(alg)
       
#         for s in alg:            
#             # if s in ['+','-','/','*']:
#             #     print("k")
#             alg.split('-')
# for x in range(9,14):
   
        
#         alg= ws[temp].value
#         print(alg)
#     print("------------")














       



    